from openpyxl import *
import os.path as op 

nosections = 2
default_section = ["Aguinaldo", "Rizal"]
directory = r'C:\\Users\\lukea\\Documents\\Desktop Junk\\School\\2020-2021\\StuGraDP\\Sheets\\'
filetype = r".xlsx"

creating = input(r'New school year:')

school_year = input(r'School year:')

filedestination = directory + school_year + filetype

file_exists = op.isfile(filedestination)

wb = Workbook()

if creating == "True":
    wb = Workbook()
    wb.save(filedestination)
    wb = load_workbook(filedestination)
    y = 0
    x = 0
    while y < nosections:
        if default_section[x] in wb.sheetnames:
            print("This section exists")
            y = y-1
        else:
            wb.create_sheet(default_section[x])
            print("Created sheet" + " " + default_section[x])
            wb.save(filename = filedestination)
        y = y+1
        x = x+1
    exit()
else:
    pass
