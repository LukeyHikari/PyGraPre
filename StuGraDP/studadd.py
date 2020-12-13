from openpyxl import *
import os.path as op 

def add_students():
    school_year = input(r'School Year:')
    nostudents = int(input(r'No. of Students:'))
    section = input(r'Section:')
    nostudents = nostudents + 1

    directory = r'C:\\Users\\lukea\\Documents\\Desktop Junk\\School\\2020-2021\\StuGraDP\\Sheets'
    filetype = r".xlsx"
    filedestination = directory + school_year + filetype
    file_exists = op.isfile(filedestination)

    wb = Workbook()

    if file_exists:
        print("Adding Students")
        wb = Workbook()
        wb = load_workbook(filedestination)
        sheet = wb[section]
        i = 1
        while i < nostudents:   
            studentname = input(r'Student Full Name:')
            print(studentname)
            confirmation = input(r'Press ENTER if the student name is right')
            if confirmation == '':
                studcell = sheet.cell(row = i, column = 1)
                studcell.value = studentname
            else:
                i = i - 1
            i = i + 1
            wb.save(filename = filedestination)
        
    else:
        print("School Year does not Exist")
