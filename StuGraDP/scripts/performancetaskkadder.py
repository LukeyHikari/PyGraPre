from openpyxl import workbook
from openpyxl import worksheet
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path as op 
import os
import csv
import pandas as pd 
import tkinter
import UI
from tkinter import Entry
from tkinter import ttk

i = 0
highestpossiblegrade = 0
studentgrade = 0
def ptaskadd(sy, sec, actno, qt):
    school_year = sy #input(r'School Year:')
    section = sec #input(r'Section:')
    activityno = actno #int(input(r'Activity Number:'))
    quarter = qt #input(r'Quarter:')
    csvquartername = ""

    app = UI.App()
    ib = app.get_inputbox()
    dp = app.get_displaybox()

    directory = os.getcwd()
    directory = directory + r'\\Sheets\\'
    filetype = r".xlsx"
    sydirectory = directory + school_year + r'\\' + section + r'\\'

    with open(sydirectory + 'Number of Students and Activities.csv', 'rt') as f:
        data = csv.DictReader(f)
        for row in data:
            noStuds = row['No. of Students']
            noStuds = int(noStuds)

    filedestination = sydirectory + section + filetype

    file_exists = op.isfile(filedestination)

    wb = Workbook()
    
    def gradeadder(self):
        global i
        global studentgrade
        studentgrade = int(ib.get())
        if i < noStuds:
            if studentgrade < highestpossiblegrade:
                if i != noStuds-1:
                    wb = Workbook()
                    wb = load_workbook(filedestination)
                    sheet = wb[quarter]
                    studcell = sheet.cell(row = i+3, column = activityno + 1)
                    studcell.value = studentgrade
                    wb.save(filename = filedestination)
                    studentname = sheet.cell(row = i+3, column = 1).value
                    ttodisplay = studentname + " " + "Grade: " + str(studentgrade)
                    dp.set(ttodisplay)
                elif i == noStuds-1:
                    wb = Workbook()
                    wb = load_workbook(filedestination)
                    sheet = wb[quarter]
                    studcell = sheet.cell(row = i+3, column = activityno + 1)
                    studcell.value = studentgrade
                    wb.save(filename = filedestination)
                    ib.unbind('<Return>')
                    studentname = sheet.cell(row = i+3, column = 1).value

#Logger
                    #Logger #put this at the end of grading funciton
                    print(r"Checking if all students are graded")
                    wb = Workbook()
                    wb = load_workbook(filedestination, read_only = True)
                    sheet = wb[quarter]

                    if quarter == 'Quarter 1':
                        csvquartername = "Q1"
                    elif quarter == 'Quarter 2':
                        csvquartername = "Q2"
                    elif quarter == 'Quarter 3':
                        csvquartername = "Q3"
                    elif quarter == 'Quarter 4':
                        csvquartername = "Q4"
                    quartercsv = sydirectory + csvquartername + r'.csv'
                    print(quartercsv)
                    for x in range(noStuds):
                        checkingcell = sheet.cell(row = x + 3, column = activityno + 1).value
                        if checkingcell > 0:
                            df = pd.read_csv(quartercsv)
                            df.loc[activityno - 1, "Performance Task Status"] = True
                            df.to_csv(quartercsv, index = False)
                            ttodisplay = studentname + " " + "Grade: " + str(studentgrade) + "\n" + "Finished Grading"
                            dp.set(ttodisplay)
                        elif checkingcell == 0:
                            print(r"A student was not graded")
                            df = pd.read_csv(quartercsv)
                            df.loc[activityno - 1, "Performance Task Status"] = False
                            df.to_csv(quartercsv, index = False)
                            ttodisplay = studentname + " " + "Grade: " + str(studentgrade) + "\n" + "Finished Grading" + "\n" + "A student was not Graded"
                            dp.set(ttodisplay)
                            break
#
            elif studentgrade > highestpossiblegrade:
                dp.set("Grade is higher than highest possible grade. Please Repeat")
                i = i-1

            i = i+1
            ib.delete(0,'end')
        else:
            ib.unbind('<Return>')


    def highestgradesetter(self):
        global highestpossiblegrade
        highestpossiblegrade = int(ib.get())
        ttodisplay = "Highest Possible Grade: " + ib.get() + "\n" + "Please proceed to adding the Grades starting w/ Student one"
        dp.set(ttodisplay)
        maxcell = sheet.cell(row = noStuds+3, column = activityno+1)
        maxcell.value = highestpossiblegrade
        wb.save(filename = filedestination)
        ib.delete(0,'end')
        ib.unbind('<Return>')
        ib.bind('<Return>', gradeadder)
        

    if file_exists:
        wb = load_workbook(filedestination)
        sheet = wb[quarter]
        dp.set("Please Input Highest Possible Grade" )
        ib.bind("<Return>", highestgradesetter)
    else:
        print("School Year does not Exist")

