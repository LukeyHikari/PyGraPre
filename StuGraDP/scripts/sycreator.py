from openpyxl import worksheet
from openpyxl import workbook
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path as op 
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import csvcreator as csvc

def sycreate(crsy, sy, sec, numstud):
    creating = crsy #input(r'New school year:')
    rvalue = ""
    def centeractnumbers(quarter):
        wb = load_workbook(filedestination)
        ws = wb[quarter]
        x = 1
        y = 2
        while x < 11:    
            cell = ws.cell(row = 2, column = y)
            cell.value = x
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
            x = x+1
            y = y+1
        x = 1
        y = 12
        while x < 11:    
            cell = ws.cell(row = 2, column = y)
            cell.value = x
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
            x = x+1
            y = y+1

    def fillsheetq1():
        wb = load_workbook(filedestination)
        quarter = r'Quarter 1'
        ws = wb[quarter]
        ws.merge_cells('B1:K1')
        ws.column_dimensions[get_column_letter(1)].width = 30
        cell = ws.cell(row = 1, column = 2)
        cell.value = 'Performance Task'
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws.merge_cells('L1:U1')
        cell = ws['L1']
        cell.value = 'Written Work'
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        cell = ws.cell(row = 2, column = 1)
        cell.value = "Student Names"
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        centeractnumbers(quarter)
        wb.save(filename = filedestination)

    def fillsheetq2():
        wb = load_workbook(filedestination)
        quarter = r'Quarter 2'
        ws = wb[quarter]
        ws.merge_cells('B1:K1')
        ws.column_dimensions[get_column_letter(1)].width = 30
        cell = ws.cell(row = 1, column = 2)
        cell.value = 'Performance Task'
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws.merge_cells('L1:U1')
        cell = ws['L1']
        cell.value = 'Written Work'
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        cell = ws.cell(row = 2, column = 1)
        cell.value = "Student Names"
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        centeractnumbers(quarter)
        wb.save(filename = filedestination)

    def fillsheetq3():
        wb = load_workbook(filedestination)
        quarter = r'Quarter 3'
        ws = wb[quarter]
        ws.merge_cells('B1:K1')
        ws.column_dimensions[get_column_letter(1)].width = 30
        cell = ws.cell(row = 1, column = 2)
        cell.value = 'Performance Task'
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws.merge_cells('L1:U1')
        cell = ws['L1']
        cell.value = 'Written Work'
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        cell = ws.cell(row = 2, column = 1)
        cell.value = "Student Names"
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        centeractnumbers(quarter)
        wb.save(filename = filedestination)

    def fillsheetq4():
        wb = load_workbook(filedestination)
        quarter = r'Quarter 4'
        ws = wb[quarter]
        ws.merge_cells('B1:K1')
        ws.column_dimensions[get_column_letter(1)].width = 30
        cell = ws.cell(row = 1, column = 2)
        cell.value = 'Performance Task'
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        ws.merge_cells('L1:U1')
        cell = ws['L1']
        cell.value = 'Written Work'
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        cell = ws.cell(row = 2, column = 1)
        cell.value = "Student Names"
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        centeractnumbers(quarter)
        wb.save(filename = filedestination)

    if creating == "True":
        school_year = sy #input(r'School year:')
        section = sec #input(r'Section:')
        noofstudents = int(numstud) #int(input(r'No. Of Students:'))
        directory = os.getcwd()
        sheetdir = directory + r'\\Sheets\\'
        sheetdirexist = op.exists(sheetdir)
        if sheetdirexist:
            print("Sheets directory exist")
        else:
            os.makedirs(directory + r'\\Sheets\\')
        directory = directory + r'\\Sheets\\' + school_year
        filetype = r".xlsx"
        file_exists = op.exists(directory)

        if file_exists:
            print("SY directory already exists")
            file_exists = op.exists(directory + r'\\' + section)
            if file_exists:
                print("Section Exists")
            else:
                print("Making Section in Already Existing School Year")
                os.makedirs(directory + r'\\' + section)     
        else:
            print("Making School Year Directory")
            os.makedirs(directory + r'\\')
            file_exists = op.exists(directory + r'\\' + section)
            if file_exists:
                print("Section Exists")
            else:
                print("Making Section in New School Year")
                os.makedirs(directory + r'\\' + section)

        sydirectory = directory + r'\\' + section + r'\\'

        filedestination = sydirectory + section + filetype

        file_exists = op.isfile(filedestination)

        if file_exists:
            print("This school year already exists")
            rvalue = "This class already exists"
        else:
            wb = Workbook()
            wb.save(filedestination)
            wb = load_workbook(filedestination)
            if "Quarter 1" in wb.sheetnames:
                print("Quarter Sheets Already Exists")
                rvalue = "Class Sheets are already made"
            else:
                wb.create_sheet("Quarter 1")
                wb.create_sheet("Quarter 2")
                wb.create_sheet("Quarter 3")
                wb.create_sheet("Quarter 4")
                print("Created the Quarterly Sheets")
                wb.save(filename = filedestination)
                csvc.qcsvcreate(school_year, section, noofstudents)
            fillsheetq1()
            fillsheetq2()
            fillsheetq3()
            fillsheetq4()
            rvalue = "Class Created w/ " + numstud + " Students"
    
    else:
        pass

    return(rvalue)

