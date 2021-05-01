from openpyxl import workbook
from openpyxl import worksheet
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path as op 
import os
import csv
import UI
from UI import App
from tkinter import ttk
from tkinter import Entry
import tkinter as tk

i = None
studentname = None

def addstuds(sy, sec):#bug fixing in process
    global i
    global studentname
    i = 0
    studentname = ""
    #Directory Variables
    school_year = sy
    section = sec
    directory = os.getcwd()
    directory = directory + r'\\Sheets\\'
    filetype = r".xlsx"
    sydirectory = directory + school_year + r'\\' + section + r'\\'

    #Retrieving Part of the UI
    app = App()
    ib = app.get_inputbox()
    dp = app.get_displaybox()

    with open(sydirectory + 'Number of Students and Activities.csv', 'rt') as f:
        data = csv.DictReader(f)
        for row in data:
            noStuds = row['No. of Students']
            noStuds = int(noStuds)

    filedestination = sydirectory + section + filetype

    file_exists = op.isfile(filedestination)

    wb = Workbook()

    def adder(self):
        global i
        global studentname
        if i < noStuds:
            if ib.get() != "":
                if i != noStuds-1:
                    studentname = ib.get()
                    print("Adding Student:", studentname)
                    wb = load_workbook(filedestination)
                    sheet = wb[r'Quarter 1']
                    studcell = sheet.cell(row = i+3, column = 1)
                    studcell.value = studentname
                    sheet = wb[r'Quarter 2']
                    studcell = sheet.cell(row = i+3, column = 1)
                    studcell.value = studentname
                    sheet = wb[r'Quarter 3']
                    studcell = sheet.cell(row = i+3, column = 1)
                    studcell.value = studentname
                    sheet = wb[r'Quarter 4']
                    studcell = sheet.cell(row = i+3, column = 1)
                    studcell.value = studentname
                    wb.save(filename = filedestination)
                    ttodisplay = "Added Student:" + " " + studentname
                    dp.set(ttodisplay)
                elif i == noStuds-1:
                    studentname = ib.get()
                    print("Adding Student:", studentname)
                    wb = load_workbook(filedestination)
                    sheet = wb[r'Quarter 1']
                    studcell = sheet.cell(row = i+3, column = 1)
                    studcell.value = studentname
                    sheet = wb[r'Quarter 2']
                    studcell = sheet.cell(row = i+3, column = 1)
                    studcell.value = studentname
                    sheet = wb[r'Quarter 3']
                    studcell = sheet.cell(row = i+3, column = 1)
                    studcell.value = studentname
                    sheet = wb[r'Quarter 4']
                    studcell = sheet.cell(row = i+3, column = 1)
                    studcell.value = studentname
                    wb.save(filename = filedestination)
                    ttodisplay = "Added Last Student:" + " " + studentname + "\n" + "Finished Adding Students for this class"
                    dp.set(ttodisplay)
                    ib.unbind('<Return>')
            elif ib.get() == studentname:
                print("Same student detected")
            else:
                print("No student name")
            i = i + 1
            ib.delete(0,'end')
        else:
            ib.unbind('<Return>')

    if file_exists:
        print("Adding Students")
        print(i)
        ib.bind('<Return>', adder)
        dp.set("Please type the students' names then press enter")
        wb = load_workbook(filedestination)
        sheet = wb[r'Quarter 1']
        maxcell = sheet.cell(row = noStuds+3, column = 1)
        maxcell.value = "Highest Possible Grade"
        sheet = wb[r'Quarter 2']
        maxcell = sheet.cell(row = noStuds+3, column = 1)
        maxcell.value = "Highest Possible Grade"
        sheet = wb[r'Quarter 3']
        maxcell = sheet.cell(row = noStuds+3, column = 1)
        maxcell.value = "Highest Possible Grade"
        sheet = wb[r'Quarter 4']
        maxcell = sheet.cell(row = noStuds+3, column = 1)
        maxcell.value = "Highest Possible Grade"
        wb.save(filename = filedestination)
        
    else:
        print("School Year does not Exist")

