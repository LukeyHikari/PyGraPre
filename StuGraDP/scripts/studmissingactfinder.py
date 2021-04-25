from openpyxl import workbook
from openpyxl import worksheet
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path as op 
import os
import csv

def findstudsmisacts(sy, sec, qt, atype, anum): #just use a return array here
    school_year = sy #input(r'School Year:')
    section = sec #input(r'Section:')
    sheetname = qt #input(r'Quarter:')
    activitytype = atype #input(r'Activity Type:')
    activityno = anum #int(input(r'Activity Number:'))
    swmacts = []
    returnvalue = None
    
    directory = os.getcwd()
    directory = directory + r'\\Sheets\\'
    filetype = r".xlsx"
    sydirectory = directory + school_year + r'\\' + section + r'\\'

    with open(sydirectory + 'Number of Students and Activities.csv', 'rt') as f:
        data = csv.DictReader(f)
        for row in data:
            noStuds = row['No. of Students']
            noStuds = int(noStuds)

    filedestination = sydirectory + section + filetype

    file_exists = op.isfile(filedestination)

    wb = Workbook()
    wb = load_workbook(filedestination, read_only = True)
    sheet = wb[sheetname]
    if file_exists:
        if activitytype == 'Performance':
                for x in range(noStuds):
                    checkingcell = sheet.cell(row = x + 3, column = activityno + 1).value
                    if checkingcell == 0:
                        studentnamecell = sheet.cell(row = x + 3, column = 1).value
                        activitynumber = str(activityno)
                        print("This student is missing Performance Activity Number" + " " + activitynumber + ":", studentnamecell)
                        swmacts.append(studentnamecell)
                        returnvalue = swmacts
                    
        elif activitytype == 'Written':
                for x in range(noStuds):
                    checkingcell = sheet.cell(row = x + 3, column = activityno + 11).value
                    if checkingcell == 0:
                        studentnamecell = sheet.cell(row = x + 3, column = 1).value
                        activitynumber = str(activityno)
                        print("This student is missing Performance Activity Number" + " " + activitynumber + ":", studentnamecell)
                        swmacts.append(studentnamecell)
                        returnvalue = swmacts
                
    else:
        print("This SY or Section does not exist")

    return(returnvalue)