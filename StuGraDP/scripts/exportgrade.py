from openpyxl import workbook
from openpyxl import worksheet
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path as op 
import os
import csv

def exportgrades(sy, sec, qt):
    school_year = sy #input(r'School Year:')
    section = sec #input(r'Section:')
    quarter = qt #(r'Quarter:')

    if quarter == '1':
        sheetname = "Quarter 1"
    elif quarter == '2':
        sheetname = "Quarter 2"
    elif quarter == '3':
        sheetname = "Quarter 3"
    elif quarter == '4':
        sheetname = "Quarter 4"


    directory = os.getcwd()
    directory = directory + r'\\Sheets\\'
    filetype = r".xlsx"
    sydirectory = directory + school_year + r'\\' + section + r'\\'

    with open(sydirectory + 'Number of Students and Activities.csv', 'rt') as f:
        data = csv.DictReader(f)
        for row in data:
            noStuds = row['No. of Students']
            noStuds = int(noStuds)

    with open(sydirectory + 'Number of Students and Activities.csv', 'rt') as f:
            data = csv.DictReader(f)
            for row in data:
                noPT = row['No. of PT']
                noPT = int(noPT)

    with open(sydirectory + 'Number of Students and Activities.csv', 'rt') as f:
            data = csv.DictReader(f)
            for row in data:
                noWT = row['No. of WT']
                noWT = int(noWT)

    filedestination = sydirectory + section + filetype

    file_exists = op.isfile(filedestination)

    wb = Workbook()
    wb = load_workbook(filedestination)
    sheet = wb[sheetname]

    noPT = noPT+1
    noWT = noWT+1
    noStuds = noStuds+1

    def getgrades():
        i = 1
        studentgradesPT = []
        studentgradesWT = []
        highestPT = []
        highestWT = []

        while i < noPT:
                gradingptcellvalue = sheet.cell(row = x+2, column = i+1).value
                studentgradesPT.append(gradingptcellvalue)
                i = i + 1
        i = 1
        while i < noWT:
                gradingwtcellvalue = sheet.cell(row = x+2, column = i+11).value
                studentgradesWT.append(gradingwtcellvalue)
                i = i + 1
        i = 1 
        while i < noPT:
                gradingptcellvalue = sheet.cell(row = noStuds+2, column = i+1).value
                highestPT.append(gradingptcellvalue)
                i = i + 1
        i = 1
        while i < noWT:
                gradingwtcellvalue = sheet.cell(row = noStuds+2, column = i+11).value
                highestWT.append(gradingwtcellvalue)
                i = i + 1

        weightedptavg = sum(studentgradesPT)/sum(highestPT)
        weightedptavg = weightedptavg * .100
        weightedptavg = weightedptavg * .50
        weightedwtavg = sum(studentgradesWT)/sum(highestWT)
        weightedwtavg = weightedwtavg * .100
        weightedwtavg = weightedwtavg * .50
        rawgrade = (weightedptavg*1000) + (weightedwtavg*1000)

        finalgrade = 0
        if rawgrade == 100:
            finalgrade = 100
        elif 98.40 <= rawgrade <= 99.99:
            finalgrade = 99
        elif 96.80 <= rawgrade <= 98.39:
            finalgrade = 98
        elif 95.20 <= rawgrade <= 96.79:
            finalgrade = 97
        elif 93.60 <= rawgrade <= 95.19:
            finalgrade = 96
        elif 92.00 <= rawgrade <= 93.59:
            finalgrade = 95
        elif 90.40 <= rawgrade <= 91.99:
            finalgrade = 94
        elif 88.80 <= rawgrade <= 90.39:
            finalgrade = 93
        elif 87.20 <= rawgrade <= 88.79:
            finalgrade = 92
        elif 85.60 <= rawgrade <= 87.19:
            finalgrade = 91
        elif 84.00 <= rawgrade <= 85.59:
            finalgrade = 90
        elif 82.40 <= rawgrade <= 83.99:
            finalgrade = 89
        elif 80.80 <= rawgrade <= 82.39:
            finalgrade = 88
        elif 79.20 <= rawgrade <= 80.79:
            finalgrade = 87
        elif 77.60 <= rawgrade <= 79.19:
            finalgrade = 86
        elif 76.00 <= rawgrade <= 77.59:
            finalgrade = 85
        elif 74.40 <= rawgrade <= 75.99:
            finalgrade = 84
        elif 72.80 <= rawgrade <= 74.39:
            finalgrade = 83
        elif 71.20 <= rawgrade <= 72.79:
            finalgrade = 82
        elif 69.60 <= rawgrade <= 71.19:
            finalgrade = 81
        elif 68.00 <= rawgrade <= 69.59:
            finalgrade = 80
        elif 66.40 <= rawgrade <= 67.99:
            finalgrade = 79
        elif 64.80 <= rawgrade <= 66.39:
            finalgrade = 78
        elif 63.20 <= rawgrade <= 64.79:
            finalgrade = 77
        elif 61.60 <= rawgrade <= 63.19:
            finalgrade = 76
        elif 60.00 <= rawgrade <= 61.59:
            finalgrade = 75
        elif 56.00 <= rawgrade <= 59.99:
            finalgrade = 74
        elif 52.00 <= rawgrade <= 55.99:
            finalgrade = 73
        elif 48.00 <= rawgrade <= 51.99:
            finalgrade = 72
        elif 44.00 <= rawgrade <= 47.99:
            finalgrade = 71
        elif 40.00 <= rawgrade <= 43.99:
            finalgrade = 70
        elif 36.00 <= rawgrade <= 39.99:
            finalgrade = 69
        elif 32.00 <= rawgrade <= 35.99:
            finalgrade = 68
        elif 28.00 <= rawgrade <= 31.99:
            finalgrade = 67
        elif 24.00 <= rawgrade <= 27.99:
            finalgrade = 66
        elif 20.00 <= rawgrade <= 23.99:
            finalgrade = 65
        elif 16.00 <= rawgrade <= 19.99:
            finalgrade = 64
        elif 12.00 <= rawgrade <= 15.99:
            finalgrade = 63
        elif 8.00 <= rawgrade <= 11.99:
            finalgrade = 62
        elif 4.00 <= rawgrade <= 7.99:
            finalgrade = 61
        elif 0 <= rawgrade <= 3.99:
            finalgrade = 60
        else:
            print("Raw grade does not exist in transmutation database")
        return finalgrade

    if file_exists:
        x = 1
        while x < noStuds:
            student = sheet.cell(row = x+2, column = 1).value
            print("Student:",student, "|", "Quarterly Grade:",getgrades())
            x = x + 1

    else:
      print("This directory does not exist")