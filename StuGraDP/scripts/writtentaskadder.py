from openpyxl import workbook
from openpyxl import worksheet
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path as op 
import os
import csv

school_year = input(r'School Year:')
section = input(r'Section:')
activityno = int(input(r'Activity Number:'))

directory = os.getcwd()
directory = directory + r'\\Sheets\\'
filetype = r".xlsx"
sydirectory = directory + school_year + r'\\' + section + r'\\'

with open(sydirectory + 'Number of Students and Activities.csv', 'rt') as f:
    data = csv.DictReader(f)
    for row in data:
        noStuds = row['No. of Students']
        noStuds = int(noStuds)

filedestination = sydirectory + section + filetype

file_exists = op.isfile(filedestination)

wb = Workbook()

if file_exists:
    print("Adding to Written Tasks")
    wb = Workbook()
    wb = load_workbook(filedestination)
    sheet = wb[section]
    i = 1
    noStuds = noStuds + 1
    while i < noStuds:   
        studentgrade= int(input(r'Student Grade:'))
        print(studentgrade)
        confirmation = input(r'Press ENTER if the grade is right')
        if confirmation == '':
            studcell = sheet.cell(row = i + 2, column = activityno + 10)
            studcell.value = studentgrade
        else:
            i = i - 1
        i = i + 1
        wb.save(filename = filedestination)
    
else:
    print("School Year does not Exist")
