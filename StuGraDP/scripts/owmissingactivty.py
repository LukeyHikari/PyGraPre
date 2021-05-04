from openpyxl import workbook
from openpyxl import worksheet
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path as op 
import os
import csv
import pandas as pd
import tkinter
import UI
from tkinter import Entry
from tkinter import ttk

def grademisact(sy, sec, qt, atype, anum, stud, agrade): #reference performance task adder(literally the same function dumbass)
    school_year = sy #input(r'School Year:')
    section = sec #input(r'Section:')
    quarter = qt #input(r'Quarter:')
    activitytype = atype #input(r'Activity Type:')
    activityno = anum #int(input(r'Activity Number:'))
    studwmisact = stud #input(r'Student with Missing Activity:')
    addedgrade = agrade
    sheetname = qt

    app = UI.App()
    ib = app.get_inputbox()
    dp = app.get_displaybox()

    directory = os.getcwd()
    directory = directory + r'\\Sheets\\'
    filetype = r".xlsx"
    sydirectory = directory + school_year + r'\\' + section + r'\\'
    
    with open(sydirectory + 'Number of Students and Activities.csv', 'rt') as f:
        data = csv.DictReader(f)
        for row in data:
            noStuds = row['No. of Students']
            noStuds = int(noStuds)

    filedestination = sydirectory + section + filetype
    file_exists = op.isfile(filedestination)

    wb = Workbook()
    wb = load_workbook(filedestination)
    sheet = wb[sheetname]
    if file_exists:
        if activitytype == 'Performance':
            highestpossiblegrade = sheet.cell(row = noStuds+3, column = activityno+1).value
            print("The highest possible grade of this activity is:", highestpossiblegrade)
            for x in range(noStuds):
                    studentcell = sheet.cell(row = x + 3, column = 1).value
                    if studentcell == studwmisact:
                        activitycell = sheet.cell(row = x + 3, column = activityno + 1)
                        studentgrade = addedgrade
                        activitycell.value = studentgrade
                        wb.save(filename = filedestination)

                        #Logger #put this at the end of grading funciton
                        print(r"Checking if all students are graded")
                        wb = Workbook()
                        wb = load_workbook(filedestination, read_only = True)
                        sheet = wb[quarter]

                        if quarter == 'Quarter 1':
                            csvquartername = "Q1"
                        elif quarter == 'Quarter 2':
                            csvquartername = "Q2"
                        elif quarter == 'Quarter 3':
                            csvquartername = "Q3"
                        elif quarter == 'Quarter 4':
                            csvquartername = "Q4"
                        quartercsv = sydirectory + csvquartername + r'.csv'
                        print(quartercsv)
                        for x in range(noStuds):
                            checkingcell = sheet.cell(row = x + 3, column = activityno + 1).value
                            if checkingcell > 0:
                                df = pd.read_csv(quartercsv)
                                df.loc[activityno - 1, "Performance Task Status"] = True
                                df.to_csv(quartercsv, index = False)
                                ttodisplay = studwmisact + " " + "Grade: " + str(studentgrade) + "\n" + "All students have been graded"
                                dp.set(ttodisplay)
                            elif checkingcell == 0:
                                print(r"A student was not graded")
                                df = pd.read_csv(quartercsv)
                                df.loc[activityno - 1, "Performance Task Status"] = False
                                df.to_csv(quartercsv, index = False)
                                ttodisplay = studwmisact + " " + "Grade: " + str(studentgrade) + "\n" + "Finished Grading" + "\n" + "A student has not been graded"
                                dp.set(ttodisplay)
                                break
                        break

        if activitytype == 'Written':
            highestpossiblegrade = sheet.cell(row = noStuds+2, column = activityno+11).value
            print("The highest possible grade of this activity is:", highestpossiblegrade)
            for x in range(noStuds):
                    studentcell = sheet.cell(row = x + 3, column = 1).value
                    if studentcell == studwmisact:
                        activitycell = sheet.cell(row = x + 3, column = activityno + 11)
                        studentgrade = addedgrade
                        activitycell.value = studentgrade
                        wb.save(filename = filedestination)

                        #Logger #put this at the end of grading funciton
                        print(r"Checking if all students are graded")
                        wb = Workbook()
                        wb = load_workbook(filedestination, read_only = True)
                        sheet = wb[quarter]

                        if quarter == 'Quarter 1':
                            csvquartername = "Q1"
                        elif quarter == 'Quarter 2':
                            csvquartername = "Q2"
                        elif quarter == 'Quarter 3':
                            csvquartername = "Q3"
                        elif quarter == 'Quarter 4':
                            csvquartername = "Q4"
                        quartercsv = sydirectory + csvquartername + r'.csv'
                        print(quartercsv)
                        for x in range(noStuds):
                            checkingcell = sheet.cell(row = x + 3, column = activityno + 11).value
                            if checkingcell > 0:
                                df = pd.read_csv(quartercsv)
                                df.loc[activityno - 1, "Written Task Status"] = True
                                df.to_csv(quartercsv, index = False)
                                ttodisplay = studwmisact + " " + "Grade: " + str(studentgrade) + "\n" + "All students have been graded"
                                dp.set(ttodisplay)
                            elif checkingcell == 0:
                                print(r"A student was not graded")
                                df = pd.read_csv(quartercsv)
                                df.loc[activityno - 1, "Written Task Status"] = False
                                df.to_csv(quartercsv, index = False)
                                ttodisplay = studwmisact + " " + "Grade: " + str(studentgrade) + "\n" + "Finished Grading" + "\n" + "A student has not  been Graded"
                                dp.set(ttodisplay)
                                break
                        break
    else:
        print("This SY or Section does not exist")